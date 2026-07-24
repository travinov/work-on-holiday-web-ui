from __future__ import annotations

import argparse
from copy import copy
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

from report_first_management import build_report_dataframe as build_report_1
from report_four_reconciliation import build_report_dataframe as build_report_4
from report_second_requests import build_report_dataframe as build_report_2
from report_third_closure import build_report_dataframe as build_report_3


def autosize_columns(worksheet) -> None:
    for column_cells in worksheet.columns:
        max_len = 0
        col_letter = column_cells[0].column_letter
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            line_len = max((len(line) for line in value.splitlines()), default=len(value))
            if line_len > max_len:
                max_len = line_len
        worksheet.column_dimensions[col_letter].width = min(max_len + 2, 80)


def apply_wrap_text_by_header(worksheet, header_row: int, header_name: str) -> None:
    target_col = None
    for cell in worksheet[header_row]:
        if cell.value == header_name:
            target_col = cell.column
            break

    if not target_col:
        return

    for row_idx in range(header_row + 1, worksheet.max_row + 1):
        worksheet.cell(row=row_idx, column=target_col).alignment = Alignment(vertical="top", wrap_text=True)


def apply_double_payment_grade_highlight_by_header(worksheet, header_row: int, header_name: str) -> None:
    target_col = None
    for cell in worksheet[header_row]:
        if cell.value == header_name:
            target_col = cell.column
            break

    if not target_col:
        return

    for row_idx in range(header_row + 1, worksheet.max_row + 1):
        cell = worksheet.cell(row=row_idx, column=target_col)
        value = "" if cell.value is None else str(cell.value)
        if "грейд 12+" in value.lower():
            font = copy(cell.font)
            font.bold = True
            font.color = "FFFF0000"
            cell.font = font


def write_sheet(writer: pd.ExcelWriter, sheet_name: str, title: str, subtitle: str, report_df: pd.DataFrame) -> None:
    report_df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=3)
    worksheet = writer.sheets[sheet_name]
    worksheet["A1"] = title
    worksheet["A2"] = subtitle

    title_font = copy(worksheet["A1"].font)
    title_font.bold = True
    title_font.size = 13
    worksheet["A1"].font = title_font

    subtitle_font = copy(worksheet["A2"].font)
    subtitle_font.italic = True
    worksheet["A2"].font = subtitle_font

    apply_wrap_text_by_header(worksheet, header_row=4, header_name="Перечень АС")
    if sheet_name == "Отчет 1":
        apply_double_payment_grade_highlight_by_header(worksheet, header_row=4, header_name="Условия выхода")
    last_column = get_column_letter(max(len(report_df.columns), 1))
    worksheet.auto_filter.ref = f"A4:{last_column}{max(worksheet.max_row, 4)}"
    worksheet.freeze_panes = "A5"
    autosize_columns(worksheet)


def build_subtitle(date_from, date_to, prepared_label: str) -> str:
    if date_from and date_to:
        period = f"{date_from.strftime('%d.%m.%Y')}–{date_to.strftime('%d.%m.%Y')}"
    else:
        period = "все даты"
    return f"Период: {period}; Дата подготовки: {prepared_label}"


def build_argument_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Сборка единого файла отчетов по выходам в выходные")
    parser.add_argument("--db", default="survey_results.db", help="Путь к SQLite БД")
    parser.add_argument("--date-from", help="Опционально: начало периода по плановой/фактической дате (YYYY-MM-DD)")
    parser.add_argument("--date-to", help="Опционально: конец периода по плановой/фактической дате (YYYY-MM-DD)")
    parser.add_argument(
        "--employees-csv",
        help="Устаревший параметр совместимости; значение игнорируется",
    )
    parser.add_argument(
        "--output-dir",
        default="reports",
        help="Папка для сохранения итогового файла",
    )
    parser.add_argument(
        "--output",
        help="Полный путь к итоговому файлу. Если не указан, используется шаблон имени по дате подготовки.",
    )
    return parser


def main() -> None:
    args = build_argument_parser().parse_args()

    db_path = Path(args.db)
    if not db_path.exists():
        raise FileNotFoundError(f"БД не найдена: {db_path}")

    prepared_at = datetime.now()
    prepared_label = prepared_at.strftime("%d.%m.%Y %H:%M")

    date_from = datetime.strptime(args.date_from, "%Y-%m-%d").date() if args.date_from else None
    date_to = datetime.strptime(args.date_to, "%Y-%m-%d").date() if args.date_to else None

    if (date_from is None) != (date_to is None):
        raise ValueError("Нужно указать обе даты: --date-from и --date-to")
    if date_from and date_to and date_from > date_to:
        raise ValueError("date-from не может быть позже date-to")
    subtitle = build_subtitle(date_from, date_to, prepared_label)

    report_1_df = build_report_1(str(db_path), date_from, date_to, employees_csv=args.employees_csv)
    report_2_df = build_report_2(str(db_path), date_from, date_to)
    report_3_df = build_report_3(str(db_path), date_from, date_to)
    report_4_df = build_report_4(str(db_path), date_from, date_to)

    if args.output:
        output_path = Path(args.output)
    else:
        filename = f"Отчеты выхода выходные {prepared_at.date().isoformat()}.xlsx"
        output_path = Path(args.output_dir) / filename

    output_path.parent.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        write_sheet(
            writer,
            sheet_name="Отчет 1",
            title="Отчет 1 для руководства (плановые заявки)",
            subtitle=subtitle,
            report_df=report_1_df,
        )
        write_sheet(
            writer,
            sheet_name="Отчет 2",
            title="Отчет 2 для заведения заявок (плановые заявки)",
            subtitle=subtitle,
            report_df=report_2_df,
        )
        write_sheet(
            writer,
            sheet_name="Отчет 3",
            title="Отчет 3 для закрытия в Пульсе (фактические выходы)",
            subtitle=subtitle,
            report_df=report_3_df,
        )
        write_sheet(
            writer,
            sheet_name="Отчет 4",
            title="Отчет 4: сверка заявившихся и предоставивших фактическое время",
            subtitle=subtitle,
            report_df=report_4_df,
        )

    print(f"Дата подготовки: {prepared_label}")
    print(f"Отчет 1 строк: {len(report_1_df)}")
    print(f"Отчет 2 строк: {len(report_2_df)}")
    print(f"Отчет 3 строк: {len(report_3_df)}")
    print(f"Отчет 4 строк: {len(report_4_df)}")
    print(f"Файл: {output_path}")


if __name__ == "__main__":
    main()

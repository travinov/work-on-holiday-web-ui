from __future__ import annotations

import argparse
import sqlite3
from contextlib import closing
from copy import copy
from datetime import date, datetime
from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment

try:
    from src.app_request_state import ensure_app_tables
    from src.work_time import LUNCH_WARNING, needs_lunch_warning
except ModuleNotFoundError:
    from app_request_state import ensure_app_tables
    from work_time import LUNCH_WARNING, needs_lunch_warning

GRADE_WARNING_TEXT = "Двойная оплата (грейд 12+)"


def format_systems_value(value: object) -> str:
    if pd.isna(value):
        return ""

    systems = [part.strip() for part in str(value).split("|") if part.strip()]
    formatted_systems = []
    for system in systems:
        formatted_systems.append(system.split("->")[-1].strip())

    return "\n".join([system for system in formatted_systems if system])


def lunch_warning_comment(value: object) -> str:
    if value is None or pd.isna(value):
        return ""
    try:
        return LUNCH_WARNING if needs_lunch_warning(str(value).strip()) else ""
    except ValueError:
        return ""


def apply_profile_validations(df: pd.DataFrame) -> pd.DataFrame:
    validated = df.copy()
    is_double_pay = (
        validated["exit_conditions"].fillna("").astype(str).str.strip().str.lower().eq("двойная оплата")
    )
    high_grade = pd.to_numeric(validated["grade_12_plus"], errors="coerce").fillna(0).eq(1)
    highlight_mask = is_double_pay & high_grade
    validated.loc[highlight_mask, "exit_conditions"] = GRADE_WARNING_TEXT
    validated["comment"] = validated["planned_work_time"].map(lunch_warning_comment)
    return validated


def build_report_dataframe(
    db_path: str,
    date_from: date | None = None,
    date_to: date | None = None,
    employees_csv: str | None = None,
) -> pd.DataFrame:
    del employees_csv  # Deprecated compatibility argument; employee data now comes only from SQLite.
    candidate_date_filter = ""
    params: list[str] = []

    if date_from and date_to:
        candidate_date_filter = "AND r.planned_work_date BETWEEN ? AND ?"
        params.extend([date_from.isoformat(), date_to.isoformat()])

    query = f"""
    WITH candidates AS (
        SELECT
            r.response_id,
            r.full_name_key,
            COALESCE(r.full_name_normalized, r.full_name) AS full_name,
            COALESCE(st.override_planned_work_date, r.planned_work_date) AS planned_work_date,
            COALESCE(st.override_planned_work_time, r.planned_work_time) AS planned_work_time,
            COALESCE(st.override_payment_type, r.payment_type) AS exit_conditions,
            COALESCE(st.override_task_description, r.task_description) AS task_description,
            COALESCE(st.override_justification, r.justification) AS justification,
            st.override_systems,
            COALESCE(st.status, 'active') AS request_status,
            COALESCE(ep.grade_12_plus, 0) AS grade_12_plus
        FROM survey_responses r
        LEFT JOIN app_request_state st ON st.response_id = r.response_id
        LEFT JOIN app_employee_profile ep ON ep.full_name_key = r.full_name_key
        WHERE r.request_type = 'Подать заявку'
          AND COALESCE(st.override_planned_work_date, r.planned_work_date) IS NOT NULL
          {candidate_date_filter.replace("r.planned_work_date", "COALESCE(st.override_planned_work_date, r.planned_work_date)")}
    ),
    actual_dates AS (
        SELECT
            r.full_name_key,
            r.actual_work_date
        FROM survey_responses r
        WHERE r.request_type = 'Указать отработанное время'
          AND r.actual_work_date IS NOT NULL
          AND r.actual_work_time IS NOT NULL
        UNION
        SELECT
            r.full_name_key,
            st.actual_work_date
        FROM app_request_state st
        JOIN survey_responses r ON r.response_id = st.response_id
        WHERE st.actual_work_date IS NOT NULL
          AND st.actual_work_time IS NOT NULL
    )
    SELECT
        c.response_id,
        c.full_name,
        (
            SELECT COUNT(*)
            FROM actual_dates ac2
            WHERE ac2.full_name_key = c.full_name_key
              AND ac2.actual_work_date BETWEEN date(c.planned_work_date, '-29 day') AND c.planned_work_date
        ) AS exits_last_month,
        c.exit_conditions,
        c.task_description,
        c.justification,
        COALESCE(c.override_systems, GROUP_CONCAT(s.system_name, ' | ')) AS systems,
        c.planned_work_date,
        c.planned_work_time,
        c.grade_12_plus
    FROM candidates c
    LEFT JOIN response_systems s ON s.response_id = c.response_id
    WHERE c.request_status IN ('active', 'in_progress', 'in_fact', 'completed')
    GROUP BY
        c.response_id,
        c.full_name,
        c.exit_conditions,
        c.task_description,
        c.justification,
        c.override_systems,
        c.planned_work_date,
        c.planned_work_time,
        c.grade_12_plus
    ORDER BY c.planned_work_date, c.full_name;
    """

    with closing(sqlite3.connect(db_path)) as conn, conn:
        ensure_app_tables(conn)
        df = pd.read_sql_query(query, conn, params=params)

    if df.empty:
        return pd.DataFrame(
            columns=[
                "ФИО",
                "Количество выходов за последний месяц",
                "Условия выхода",
                "Перечень задач",
                "Обоснование привлечения",
                "Перечень АС",
                "Плановая дата выхода",
                "Плановое время работ",
                "Комментарий",
            ]
        )

    df["planned_work_date"] = pd.to_datetime(df["planned_work_date"], errors="coerce").dt.strftime("%d.%m.%Y")
    df["exits_last_month"] = pd.to_numeric(df["exits_last_month"], errors="coerce").fillna(0).astype(int)
    df["systems"] = df["systems"].apply(format_systems_value)
    df = apply_profile_validations(df)

    report_df = df.rename(
        columns={
            "full_name": "ФИО",
            "exits_last_month": "Количество выходов за последний месяц",
            "exit_conditions": "Условия выхода",
            "task_description": "Перечень задач",
            "justification": "Обоснование привлечения",
            "systems": "Перечень АС",
            "planned_work_date": "Плановая дата выхода",
            "planned_work_time": "Плановое время работ",
            "comment": "Комментарий",
        }
    )

    columns = [
        "ФИО",
        "Количество выходов за последний месяц",
        "Условия выхода",
        "Перечень задач",
        "Обоснование привлечения",
        "Перечень АС",
        "Плановая дата выхода",
        "Плановое время работ",
        "Комментарий",
    ]

    return report_df[columns]


def find_column_index(worksheet, header_row: int, header_name: str) -> int | None:
    for cell in worksheet[header_row]:
        if cell.value == header_name:
            return cell.column
    return None


def apply_report_styles(worksheet, header_row: int) -> None:
    systems_col = find_column_index(worksheet, header_row, "Перечень АС")
    if systems_col:
        for row_idx in range(header_row + 1, worksheet.max_row + 1):
            worksheet.cell(row=row_idx, column=systems_col).alignment = Alignment(vertical="top", wrap_text=True)

    conditions_col = find_column_index(worksheet, header_row, "Условия выхода")
    if conditions_col:
        for row_idx in range(header_row + 1, worksheet.max_row + 1):
            cell = worksheet.cell(row=row_idx, column=conditions_col)
            value = "" if cell.value is None else str(cell.value)
            if "грейд 12+" in value.lower():
                font = copy(cell.font)
                font.bold = True
                font.color = "FFFF0000"
                cell.font = font


def save_report(report_df: pd.DataFrame, output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        report_df.to_excel(writer, sheet_name="Отчет 1", index=False)
        worksheet = writer.sheets["Отчет 1"]
        apply_report_styles(worksheet, header_row=1)


def parse_date(value: str) -> date:
    return datetime.strptime(value, "%Y-%m-%d").date()


def build_argument_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Формирование отчета №1 для руководства")
    parser.add_argument("--db", default="survey_results.db", help="Путь к SQLite БД")
    parser.add_argument("--date-from", help="Опционально: начало периода по плановой дате (YYYY-MM-DD)")
    parser.add_argument("--date-to", help="Опционально: конец периода по плановой дате (YYYY-MM-DD)")
    parser.add_argument(
        "--employees-csv",
        help="Устаревший параметр совместимости; значение игнорируется",
    )
    parser.add_argument(
        "--output",
        help="Путь для xlsx-отчета. По умолчанию: reports/management_report_1_planned.xlsx",
    )
    return parser


def main() -> None:
    args = build_argument_parser().parse_args()

    db_path = Path(args.db)
    if not db_path.exists():
        raise FileNotFoundError(f"БД не найдена: {db_path}")

    date_from = parse_date(args.date_from) if args.date_from else None
    date_to = parse_date(args.date_to) if args.date_to else None

    if (date_from is None) != (date_to is None):
        raise ValueError("Нужно указать обе даты: --date-from и --date-to")
    if date_from and date_to and date_from > date_to:
        raise ValueError("date-from не может быть позже date-to")

    default_output = Path("reports") / "management_report_1_planned.xlsx"
    output_path = Path(args.output) if args.output else default_output

    report_df = build_report_dataframe(str(db_path), date_from, date_to, args.employees_csv)
    save_report(report_df, output_path)

    if date_from and date_to:
        print(f"Фильтр по плановой дате: {date_from.isoformat()} .. {date_to.isoformat()}")
    else:
        print("Фильтр по дате: не задан (все неотмененные заявки)")
    print(f"Строк в отчете: {len(report_df)}")
    print(f"Файл: {output_path}")


if __name__ == "__main__":
    main()
